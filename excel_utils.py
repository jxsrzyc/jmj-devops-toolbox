"""Excel 工具模块 - 导出和模板生成"""

import io
from datetime import datetime
from flask import send_file


def make_excel(headers, rows, sheet_name="Sheet1"):
    """生成 xlsx 文件 (内存中)"""
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    # 表头加粗
    from openpyxl.styles import Font, PatternFill, Alignment
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    center = Alignment(horizontal="center", vertical="center")

    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center

    # 数据行
    for row_idx, row in enumerate(rows, 2):
        for col_idx, v in enumerate(row, 1):
            ws.cell(row=row_idx, column=col_idx, value=v)

    # 自动列宽
    for col_idx, h in enumerate(headers, 1):
        col_letter = ws.cell(row=1, column=col_idx).column_letter
        max_len = max([len(str(h))] + [len(str(r[col_idx-1])) for r in rows if r] + [10])
        ws.column_dimensions[col_letter].width = min(max_len * 1.5, 50)

    # 冻结首行
    ws.freeze_panes = "A2"

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out


def excel_response(headers, rows, filename, sheet_name="Sheet1"):
    """生成 xlsx 响应"""
    bio = make_excel(headers, rows, sheet_name)
    return send_file(
        bio,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename,
    )


def format_dt(value):
    """格式化日期时间"""
    if not value:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    return str(value)